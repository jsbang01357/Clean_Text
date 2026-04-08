#!/usr/bin/env python3
"""
lab_to_table.py

EMR / lab raw text -> Streamlit table + Excel export + TSV converter

핵심 기능
- raw lab text에서 검사 결과 row 추출
- 검사명 정리: (응급), (응급뇨), 앞쪽 들여쓰기 기호(‥, .. 등) 제거
- 이상표시(▲/▼)는 결과값 오른쪽에 붙임
- 여러 검사 블록을 section + 채혈일 기준으로 그룹화
- Streamlit UI에서 표 미리보기 + Excel 다운로드
- CLI에서는 TSV 출력 유지
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
import datetime
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

# ===== Parser constants =====

SKIP_EXACT = {
    "[소견]",
    "[의뢰의사 Comment]",
    "Antibiotic",
}

SECTION_HEADER_RE = re.compile(r"^\s*\[진검\]\s*(.+?)\s*$")
META_LINE_RE = re.compile(r"(채혈:|접수:|보고:|IMN\()")
COLUMN_HEADER_RE = re.compile(r"^\s*검사명\s+결과값\s+단위\s+참고치\s*$")
COMMENT_LINE_RE = re.compile(r"^\s*(\[소견\]|\[의뢰의사 Comment\]|\[판독의\])")
MARKDOWN_HEADER_RE = re.compile(r"^\s*\|\s*Test\s*\|\s*Value\s*\|", re.I)
MARKDOWN_RULE_RE = re.compile(r"^\s*\|\s*-{2,}")
BULLET_PREFIX_RE = re.compile(r"^[\s\.·‥∙⋅•ㆍ]+")
EMERGENCY_PREFIX_RE = re.compile(r"^\((?:응급|응급뇨)\)\s*")
DRAW_TIME_RE = re.compile(r"채혈:\s*(\d{4}-\d{2}-\d{2})\s*(\d{2}:\d{2})")

UNIT_ALIASES = {
    "㎕": "uL",
    "μL": "uL",
    "µL": "uL",
    "μIU": "uIU",
    "µIU": "uIU",
    "㎗": "dL",
}


def normalize_unit_token(s: str) -> str:
    s = s.strip()
    for src, dst in UNIT_ALIASES.items():
        s = s.replace(src, dst)
    s = re.sub(r"/dl\b", "/dL", s, flags=re.I)
    s = re.sub(r"/ml\b", "/mL", s, flags=re.I)
    s = re.sub(r"/ul\b", "/uL", s, flags=re.I)
    s = re.sub(r"\bmg/dl\b", "mg/dL", s, flags=re.I)
    s = re.sub(r"\bg/dl\b", "g/dL", s, flags=re.I)
    s = re.sub(r"\bug/ml\b", "ug/mL", s, flags=re.I)
    s = re.sub(r"\bng/ml\b", "ng/mL", s, flags=re.I)
    s = re.sub(r"\bpg/ml\b", "pg/mL", s, flags=re.I)
    s = re.sub(r"\bug/dl\b", "ug/dL", s, flags=re.I)
    return s


SPECIAL_UNIT_PATTERNS = [
    r"mL/min/1\.73m²",
    r"mOsm/kg H2O",
    r"mmHg",
    r"mm/h",
    r"sec",
    r"fL",
    r"pg",
    r"%",
]

COUNT_UNIT_PATTERNS = [
    r"×10\^?\d+/(?:uL|[^\s]+)",
    r"×10³/(?:uL|[^\s]+)",
    r"×10\^6/(?:uL|[^\s]+)",
    r"cells/HPF",
    r"/LPF",
]

CONCENTRATION_UNIT_PATTERNS = [
    r"(?:mg|g|ug|ng|pg)/(?:dL|L|mL)",
    r"(?:IU|U|uIU)/(?:L|mL)",
    r"mmol/L",
    r"mEq/L",
    r"g/L",
    r"mg/mL",
]

RATIO_UNIT_PATTERNS = [
    r"mmol/mol",
    r"mL/dL",
]

UNIT_PATTERNS = (
    SPECIAL_UNIT_PATTERNS
    + COUNT_UNIT_PATTERNS
    + CONCENTRATION_UNIT_PATTERNS
    + RATIO_UNIT_PATTERNS
)
UNIT_RE = re.compile(
    r"^(?:" + "|".join(f"(?:{p})" for p in UNIT_PATTERNS) + r")$",
    re.I,
)
UNIT_REF_SPLIT_RE = re.compile(
    r"^(?P<unit>(?:" + "|".join(f"(?:{p})" for p in UNIT_PATTERNS) + r"))\s+(?P<ref>.+)$",
    re.I,
)
FLAG_RE = re.compile(r"^[▲▼]$")

QUAL_SECTION_KEYWORDS = ("미생물", "혈액은행", "병리", "영상")
QUAL_SKIP_EXACT = {"Gram Stain & Cul & Sensi", "Blood culture", "Antibiotic"}
RANGE_HINT_RE = re.compile(r"[~<>≤≥]|\b경계치\b|\b높음\b|\b정상\b|\b이상지질혈증\b|\b신부전\b|\b감소\b")
CONTINUATION_REF_START_RE = re.compile(
    r"""^\s*(
        [<>≤≥]\s*\d
        |
        \d+(?:\.\d+)?\s*~\s*\d+(?:\.\d+)?
        |
        \d+(?:\.\d+)?\s*이상
        |
        \d+(?:\.\d+)?\s*이하
    )""",
    re.X,
)
CONTINUATION_REF_KEYWORD_RE = re.compile(
    r"(경계치|높음|낮음|정상|이상지질혈증|대사증후군|신부전|경도감소|고도감소)"
)
RANGE_LIKE_NAME_RE = re.compile(
    r"""^\s*(
        [<>≤≥]\s*\d
        |
        \d+(?:\.\d+)?\s*~\s*\d+(?:\.\d+)?
    )""",
    re.X,
)

SECTION_TITLE_BLACKLIST = {
    "CBC with diff count & ESR",
    "WBC differential count",
    "Admission/Electro Battery(24",
    "종)",
    "Routine U/A with Microscope",
    "Gram Stain & Cul & Sensi",
    "Urine Microscopy",
}


# ===== Data model =====

@dataclass
class LabRow:
    name: str
    value: str
    unit: str = ""
    ref: str = ""
    raw_line: str = ""
    section: str = ""
    draw_time: str = ""
    table_title: str = ""

@dataclass
class QualRow:
    item: str
    result: str
    note: str = ""
    raw_line: str = ""
    section: str = ""
    draw_time: str = ""
    table_title: str = ""


# ===== Utility =====

def normalize_line(line: str) -> str:
    line = line.replace("\u3000", " ")
    line = line.replace("\t", " ")
    line = line.rstrip()
    line = re.sub(r" {2,}", "  ", line)
    return line


def split_columns(line: str) -> List[str]:
    parts = [p.strip() for p in re.split(r" {2,}", line.strip())]
    return [p for p in parts if p]


def clean_test_name(name: str) -> str:
    name = BULLET_PREFIX_RE.sub("", name)
    while True:
        new = EMERGENCY_PREFIX_RE.sub("", name)
        new = BULLET_PREFIX_RE.sub("", new)
        if new == name:
            break
        name = new
    name = re.sub(r"\s+", " ", name).strip()
    return name


def clean_section_name(section: str) -> str:
    section = section.strip()
    section = re.sub(r"\[[^\]]+\]", "", section)
    section = re.sub(r"\s+", " ", section).strip(" -")
    return section


def extract_section_name(line: str) -> str:
    m = SECTION_HEADER_RE.match(line.strip())
    if not m:
        return "Unknown"
    return clean_section_name(m.group(1)) or "Unknown"


def extract_draw_time(line: str) -> str:
    m = DRAW_TIME_RE.search(line)
    if not m:
        return ""
    return f"{m.group(1)} {m.group(2)}"


def format_title(section: str, draw_time: str) -> str:
    if draw_time:
        try:
            dt = datetime.strptime(draw_time, "%Y-%m-%d %H:%M")
            return f"{section}_{dt.year}_{dt.month:02d}_{dt.day:02d}"
        except ValueError:
            pass
    return section


def is_qualitative_section(section: str) -> bool:
    return any(k in (section or "") for k in QUAL_SECTION_KEYWORDS)


def is_qual_skip_line(line: str) -> bool:
    s = line.strip()
    return (
        not s
        or s in QUAL_SKIP_EXACT
        or COLUMN_HEADER_RE.match(s) is not None
        or COMMENT_LINE_RE.match(s) is not None
        or META_LINE_RE.search(s) is not None
    )


def is_section_title_like(line: str) -> bool:
    s = clean_test_name(line)
    return s in SECTION_TITLE_BLACKLIST


def is_bare_range_line(line: str) -> bool:
    s = line.strip()
    return bool(RANGE_LIKE_NAME_RE.match(s))


def is_qualitative_continuation(line: str) -> bool:
    s = line.strip()
    if not s or is_skip_line(s):
        return False
    if is_bare_range_line(s):
        return False
    parts = split_columns(s)
    if len(parts) == 1:
        return True
    if len(parts) == 2 and not parse_candidate_row(s):
        return True
    return False


def is_skip_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    if s in SKIP_EXACT:
        return True
    if SECTION_HEADER_RE.match(s):
        return True
    if COLUMN_HEADER_RE.match(s):
        return True
    if COMMENT_LINE_RE.match(s):
        return True
    if MARKDOWN_HEADER_RE.match(s) or MARKDOWN_RULE_RE.match(s):
        return True
    if META_LINE_RE.search(s):
        return True
    if s.startswith("- "):
        return True
    return False


def looks_like_continuation_ref(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if is_skip_line(s):
        return False

    # 실제 검사 row처럼 보이는 줄은 continuation으로 처리하지 않음.
    # 예: "RBC 2.65 ▼ ×10^6/㎕ 4.2~6.3"
    parts = split_columns(s)
    if len(parts) >= 3 and not RANGE_LIKE_NAME_RE.match(parts[0]):
        return False

    if CONTINUATION_REF_START_RE.search(s):
        return True
    if CONTINUATION_REF_KEYWORD_RE.search(s) and len(s) < 120:
        return True
    if s.startswith("Total cholesterol -"):
        return True
    return False


def append_ref(old_ref: str, extra: str) -> str:
    extra = re.sub(r"\s+", " ", extra).strip()
    if not old_ref:
        return extra
    old_ref = old_ref.strip()
    if old_ref.endswith((";", ",")):
        return f"{old_ref} {extra}"
    return f"{old_ref}; {extra}"


def compose_value(value: str, flag: str = "") -> str:
    value = value.strip()
    flag = flag.strip()
    if value and flag:
        return f"{value} {flag}"
    return value or flag


def value_flag(value: str) -> str:
    if value.endswith("▲"):
        return "up"
    if value.endswith("▼"):
        return "down"
    return ""


# ===== Parsing =====

def parse_candidate_row(line: str, section: str = "", draw_time: str = "") -> Optional[LabRow]:
    line = normalize_line(line)
    if RANGE_LIKE_NAME_RE.match(line):
        return None

    parts = split_columns(line)
    if len(parts) < 2:
        return None

    name = ""
    value = ""
    flag = ""
    unit = ""
    ref = ""

    if len(parts) >= 4:
        name = parts[0]
        value = parts[1]
        idx = 2

        if idx < len(parts) and FLAG_RE.match(parts[idx]):
            flag = parts[idx]
            idx += 1

        if idx < len(parts):
            part_norm = normalize_unit_token(parts[idx])
            if UNIT_RE.match(part_norm):
                unit = part_norm
                idx += 1

        if idx < len(parts):
            ref = " ".join(parts[idx:]).strip()

        if not unit and ref:
            ref_norm = normalize_unit_token(ref)
            if UNIT_RE.match(ref_norm):
                unit, ref = ref_norm, ""

    elif len(parts) == 3:
        name, second, third = parts
        m = re.match(r"^(.*?)(?:\s*([▲▼]))?$", second)
        if m:
            value = m.group(1).strip()
            flag = (m.group(2) or "").strip()

        third = third.strip()
        third_norm = normalize_unit_token(third)
        if UNIT_RE.match(third_norm):
            unit = third_norm
        else:
            m_ur = UNIT_REF_SPLIT_RE.match(third_norm)
            if m_ur:
                unit = m_ur.group("unit").strip()
                ref = m_ur.group("ref").strip()
            else:
                ref = third

    else:  # len(parts) == 2
        name, second = parts
        m = re.match(r"^(.*?)(?:\s*([▲▼]))?$", second)
        if m:
            value = m.group(1).strip()
            flag = (m.group(2) or "").strip()

    name = clean_test_name(name)

    if not name:
        return None
    if name in SECTION_TITLE_BLACKLIST:
        return None

    m2 = re.match(r"^(.*?)(?:\s*([▲▼]))$", value)
    if m2 and not flag:
        value = m2.group(1).strip()
        flag = m2.group(2)

    if name in {"검사명", "Test"} or value in {"결과값", "Value"}:
        return None
    if not value:
        return None

    value = re.sub(r"\s+", " ", value).strip()
    unit = normalize_unit_token(re.sub(r"\s+", " ", unit).strip()) if unit else ""
    ref = re.sub(r"\s+", " ", ref).strip()

    return LabRow(
        name=name,
        value=compose_value(value, flag),
        unit=unit,
        ref=ref,
        raw_line=line,
        section=section,
        draw_time=draw_time,
        table_title=format_title(section or "Unknown", draw_time),
    )


def parse_qualitative_row(line: str, section: str = "", draw_time: str = "") -> Optional[QualRow]:
    line = normalize_line(line)
    if not line or is_skip_line(line) or is_qual_skip_line(line):
        return None
    if is_section_title_like(line):
        return None

    parts = split_columns(line)
    if len(parts) >= 3:
        item = clean_test_name(parts[0])
        result = parts[1].strip()
        note = " ".join(parts[2:]).strip()
    elif len(parts) == 2:
        item = clean_test_name(parts[0])
        result = parts[1].strip()
        note = ""
    else:
        return None

    if not item or item in {"검사명", "Test"}:
        return None

    return QualRow(
        item=item,
        result=result,
        note=re.sub(r"\s+", " ", note).strip(),
        raw_line=line,
        section=section,
        draw_time=draw_time,
        table_title=format_title(section or "Unknown", draw_time),
    )


def parse_lab_text(text: str) -> Tuple[List[LabRow], List[QualRow], List[str]]:
    rows: List[LabRow] = []
    qual_rows: List[QualRow] = []
    unparsed_lines: List[str] = []
    last_row: Optional[LabRow] = None
    last_qual_row: Optional[QualRow] = None
    in_comment_block = False
    current_section = "Unknown"
    current_draw_time = ""

    for raw in text.splitlines():
        line = normalize_line(raw)
        stripped = line.strip()

        if not stripped:
            continue

        if SECTION_HEADER_RE.match(stripped):
            current_section = extract_section_name(stripped)
            current_draw_time = ""
            in_comment_block = False
            last_row = None
            last_qual_row = None
            continue

        draw_time = extract_draw_time(stripped)
        if draw_time:
            current_draw_time = draw_time
            continue

        if is_skip_line(stripped):
            if stripped in {"[소견]", "[의뢰의사 Comment]"}:
                in_comment_block = True
            continue

        if in_comment_block:
            if SECTION_HEADER_RE.match(stripped):
                in_comment_block = False
            else:
                continue

        if is_qualitative_section(current_section):
            qrow = parse_qualitative_row(stripped, section=current_section, draw_time=current_draw_time)
            if qrow:
                qual_rows.append(qrow)
                last_qual_row = qrow
                last_row = None
                continue

            if last_qual_row and is_qualitative_continuation(stripped):
                extra = re.sub(r"\s+", " ", stripped).strip()
                last_qual_row.note = append_ref(last_qual_row.note, extra)
                continue

            if stripped not in SECTION_TITLE_BLACKLIST and not is_qual_skip_line(stripped):
                unparsed_lines.append(stripped)
            continue

        if last_row and looks_like_continuation_ref(stripped):
            last_row.ref = append_ref(last_row.ref, stripped)
            continue

        row = parse_candidate_row(stripped, section=current_section, draw_time=current_draw_time)
        if row:
            rows.append(row)
            last_row = row
            last_qual_row = None
            continue

        if stripped not in SECTION_TITLE_BLACKLIST:
            unparsed_lines.append(stripped)

    return rows, qual_rows, unparsed_lines


# ===== Export helpers =====

def rows_to_tsv(rows: Iterable[LabRow], include_header: bool = True) -> str:
    out_lines: List[str] = []
    if include_header:
        out_lines.append("\t".join(["제목", "검사명", "결과값", "단위", "참고치"]))
    for row in rows:
        out_lines.append("\t".join([row.table_title, row.name, row.value, row.unit, row.ref]))
    return "\n".join(out_lines)


def rows_to_dataframe(rows: List[LabRow]):
    import pandas as pd

    return pd.DataFrame(
        [
            {
                "제목": r.table_title,
                "검사명": r.name,
                "결과값": r.value,
                "단위": r.unit,
                "참고치": r.ref,
                "검사종류": r.section,
                "채혈시각": r.draw_time,
            }
            for r in rows
        ]
    )


def qual_rows_to_dataframe(rows: List[QualRow]):
    import pandas as pd

    return pd.DataFrame(
        [
            {
                "제목": r.table_title,
                "항목": r.item,
                "결과": r.result,
                "비고": r.note,
                "검사종류": r.section,
                "채혈시각": r.draw_time,
            }
            for r in rows
        ]
    )


def rows_grouped(rows: List[LabRow]) -> Dict[str, List[LabRow]]:
    grouped: Dict[str, List[LabRow]] = {}
    for row in rows:
        grouped.setdefault(row.table_title or "Unknown", []).append(row)
    return grouped


def qual_rows_grouped(rows: List[QualRow]) -> Dict[str, List[QualRow]]:
    grouped: Dict[str, List[QualRow]] = {}
    for row in rows:
        grouped.setdefault(row.table_title or "Unknown", []).append(row)
    return grouped


def build_excel_bytes(rows: List[LabRow], qual_rows: List[QualRow], raw_text: str, unparsed_lines: List[str]) -> BytesIO:
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    parsed_df = rows_to_dataframe(rows)
    qual_df = qual_rows_to_dataframe(qual_rows)
    raw_df = pd.DataFrame({"원문": raw_text.splitlines()})
    unparsed_df = pd.DataFrame({"파싱되지 않은 줄": unparsed_lines})
    grouped = rows_grouped(rows)
    qual_grouped = qual_rows_grouped(qual_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        parsed_df.to_excel(writer, index=False, sheet_name="Parsed_Labs")
        qual_df.to_excel(writer, index=False, sheet_name="Qualitative_Results")
        raw_df.to_excel(writer, index=False, sheet_name="Original_Text")
        unparsed_df.to_excel(writer, index=False, sheet_name="Unparsed")

        wb = writer.book
        ws_group = wb.create_sheet("Grouped_Tables", 0)
        ws_qual_group = wb.create_sheet("Grouped_Qualitative", 1)

        title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        red_font = Font(color="C62828")
        blue_font = Font(color="1565C0")

        current_row = 1
        current_row_q = 1
        headers = ["검사명", "결과값", "단위", "참고치"]

        for title, group_rows in grouped.items():
            ws_group.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws_group.cell(row=current_row, column=1, value=title)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal="left")
            current_row += 1

            for idx, header in enumerate(headers, start=1):
                c = ws_group.cell(row=current_row, column=idx, value=header)
                c.font = header_font
                c.fill = header_fill
            current_row += 1

            for r in group_rows:
                c1 = ws_group.cell(row=current_row, column=1, value=r.name)
                c2 = ws_group.cell(row=current_row, column=2, value=r.value)
                c3 = ws_group.cell(row=current_row, column=3, value=r.unit)
                c4 = ws_group.cell(row=current_row, column=4, value=r.ref)

                flag = value_flag(r.value)
                if flag == "up":
                    for c in [c1, c2, c3, c4]:
                        c.font = red_font
                elif flag == "down":
                    for c in [c1, c2, c3, c4]:
                        c.font = blue_font
                current_row += 1

            current_row += 2

        for title, group_rows in qual_grouped.items():
            ws_qual_group.merge_cells(start_row=current_row_q, start_column=1, end_row=current_row_q, end_column=3)
            cell = ws_qual_group.cell(row=current_row_q, column=1, value=title)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal="left")
            current_row_q += 1

            for idx, header in enumerate(["항목", "결과", "비고"], start=1):
                c = ws_qual_group.cell(row=current_row_q, column=idx, value=header)
                c.font = header_font
                c.fill = header_fill
            current_row_q += 1

            for r in group_rows:
                ws_qual_group.cell(row=current_row_q, column=1, value=r.item)
                ws_qual_group.cell(row=current_row_q, column=2, value=r.result)
                ws_qual_group.cell(row=current_row_q, column=3, value=r.note)
                current_row_q += 1

            current_row_q += 2

        for sheet_name, df in {
            "Parsed_Labs": parsed_df,
            "Qualitative_Results": qual_df,
            "Original_Text": raw_df,
            "Unparsed": unparsed_df,
        }.items():
            ws = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, start=1):
                max_len = max([len(str(col))] + [len(str(v)) for v in df[col].fillna("")]) if not df.empty else len(str(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)

            if sheet_name == "Parsed_Labs" and not df.empty:
                for row_idx in range(2, len(df) + 2):
                    # 결과값(3번 컬럼)을 기준으로 플래그 확인
                    val = str(ws.cell(row=row_idx, column=3).value or "")
                    flag = value_flag(val)
                    if flag:
                        target_font = red_font if flag == "up" else blue_font
                        # 제목~채혈시각(1~7번 컬럼) 전체 색상 적용
                        for col_idx in range(1, 8):
                            ws.cell(row=row_idx, column=col_idx).font = target_font

        for idx, width in {1: 28, 2: 16, 3: 14, 4: 50}.items():
            ws_group.column_dimensions[get_column_letter(idx)].width = width
        for idx, width in {1: 28, 2: 18, 3: 60}.items():
            ws_qual_group.column_dimensions[get_column_letter(idx)].width = width

    output.seek(0)
    return output


# ===== Streamlit UI =====

def render_lab_to_excel_tool() -> None:
    import streamlit as st

    st.title("🧪 Lab → Excel 변환기")
    st.caption("검사 결과 행을 표로 정리한 뒤, 엑셀 파일로 바로 다운로드합니다.")
    st.info(
        "원문에서 검사명 / 결과값 / 단위 / 참고치를 추출합니다. "
        "검사명 앞의 (응급), (응급뇨), ‥ 같은 표시는 제거하고, ▲/▼는 결과값 오른쪽에 붙입니다. "
        "검사 블록이 바뀌면 엑셀 Grouped_Tables 시트에서 별도 표로 분리됩니다."
    )

    default_sample = """[진검]  응급혈액[WB, EDTA]\n　채혈: 2026-04-05 17:48  접수: 2026-04-06 13:45  IMN(조영일)  보고: 2026-04-06 14:04  -\n　검사명                               결과값       단위         참고치\n　　(응급)WBC                          9.13         ×10³/㎕    4~10\n　　(응급)RBC                          3.27  ▼     ×10^6/㎕    4.2~6.3\n　　‥(응급)RDW                        15.9  ▲     %            11.5~14.5\n[진검]  응급화학[Plasma, PST]\n　채혈: 2026-04-05 17:48  접수: 2026-04-06 14:11  IMN(조영일)  보고: 2026-04-06 14:28  -\n　검사명                               결과값       단위         참고치\n　　(응급)Glucose                      177  ▲      mg/dl        70~99\n　　(응급)Albumin                      2.6  ▼      g/dl         3.8~5.3\n"""

    def set_sample():
        st.session_state["lab_excel_input"] = default_sample

    def clear_input():
        st.session_state["lab_excel_input"] = ""
        st.session_state["lab_results"] = None

    with st.expander("예시 데이터 넣기", expanded=False):
        st.button("샘플 데이터 입력", on_click=set_sample, use_container_width=True)

    raw_text = st.text_area(
        "Lab / EMR 원문 붙여넣기",
        key="lab_excel_input",
        height=360,
        placeholder="여기에 검사 결과 원문을 붙여넣으세요...",
    )

    col1, col2 = st.columns([1, 1])
    with col1:
        run = st.button("표 만들기", type="primary", use_container_width=True)
    with col2:
        st.button("입력 지우기", on_click=clear_input, use_container_width=True)

    # --- 데이터 처리 로직 (버튼 클릭 시에만 실행) ---
    if run:
        if not raw_text.strip():
            st.warning("원문을 먼저 붙여넣어 주세요.")
            st.session_state["lab_results"] = None
        else:
            rows, qual_rows, unparsed_lines = parse_lab_text(raw_text)
            df = rows_to_dataframe(rows)
            qual_df = qual_rows_to_dataframe(qual_rows)
            excel_bytes = build_excel_bytes(rows, qual_rows, raw_text, unparsed_lines)
            tsv_text = rows_to_tsv(rows)
            grouped = rows_grouped(rows)
            qual_grouped = qual_rows_grouped(qual_rows)
            
            # 파싱 결과를 세션 상태에 저장하여 리런 시에도 유지
            st.session_state["lab_results"] = {
                "rows": rows,
                "qual_rows": qual_rows,
                "unparsed_lines": unparsed_lines,
                "df": df,
                "qual_df": qual_df,
                "excel_bytes": excel_bytes,
                "tsv_text": tsv_text,
                "grouped": grouped,
                "qual_grouped": qual_grouped,
            }

    # --- 결과 렌더링 로직 (세션에 결과가 있을 때 항상 실행) ---
    if "lab_results" in st.session_state and st.session_state["lab_results"] is not None:
        res = st.session_state["lab_results"]
        rows = res["rows"]
        qual_rows = res.get("qual_rows", [])
        unparsed_lines = res["unparsed_lines"]
        df = res["df"]
        qual_df = res.get("qual_df")
        excel_bytes = res["excel_bytes"]
        tsv_text = res["tsv_text"]
        grouped = res["grouped"]
        qual_grouped = res.get("qual_grouped", {})

        st.success(f"정량 {len(rows)}개 row, 정성 {len(qual_rows)}개 row를 추출했습니다.")
        metric_cols = st.columns(5)
        metric_cols[0].metric("정량 row", len(rows))
        metric_cols[1].metric("정성 row", len(qual_rows))
        metric_cols[2].metric("정량 표 수", len(grouped))
        metric_cols[3].metric("미파싱 줄", len(unparsed_lines))
        metric_cols[4].metric("원문 줄 수", len(raw_text.splitlines()))

        st.markdown("### 표 미리보기")
        for title, group_rows in grouped.items():
            st.markdown(f"#### {title}")
            group_df = rows_to_dataframe(group_rows)[["검사명", "결과값", "단위", "참고치"]]
            st.dataframe(group_df, use_container_width=True, hide_index=True)
            st.write("")

        if qual_grouped:
            st.markdown("### 정성 결과 미리보기")
            for title, group_rows in qual_grouped.items():
                st.markdown(f"#### {title}")
                group_df = qual_rows_to_dataframe(group_rows)[["항목", "결과", "비고"]]
                st.dataframe(group_df, use_container_width=True, hide_index=True)
                st.write("")

        # 파일명 생성 (KST 기준)
        kst = timezone(timedelta(hours=9))
        now_str = datetime.now(kst).strftime("%Y%m%d_%H%M")
        file_base = f"table_{now_str}"

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                label="📥 엑셀 다운로드",
                data=excel_bytes,
                file_name=f"{file_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl_col2:
            st.download_button(
                label="📥 TSV 다운로드",
                data=tsv_text,
                file_name=f"{file_base}.tsv",
                mime="text/tab-separated-values",
                use_container_width=True,
            )

        with st.expander("전체 정량 row 테이블 보기", expanded=False):
            st.dataframe(df, use_container_width=True, hide_index=True)

        if qual_df is not None and not qual_df.empty:
            with st.expander("전체 정성 row 테이블 보기", expanded=False):
                st.dataframe(qual_df, use_container_width=True, hide_index=True)

        if unparsed_lines:
            with st.expander("파싱되지 않은 줄 보기", expanded=False):
                st.code("\n".join(unparsed_lines), language="text")


# ===== CLI =====

def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Raw EMR/lab text -> Excel-friendly TSV")
    p.add_argument("input", nargs="?", help="input text file path. If omitted, read from stdin.")
    p.add_argument("-o", "--output", help="output TSV file path. If omitted, print to stdout.")
    p.add_argument("--no-header", action="store_true", help="omit TSV header row")
    return p


def read_text(path: Optional[str]) -> str:
    if path:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return sys.stdin.read()


def write_text(path: Optional[str], text: str) -> None:
    if path:
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write(text)
    else:
        sys.stdout.write(text)


def main() -> None:
    args = build_argparser().parse_args()
    raw = read_text(args.input)
    rows, _, _ = parse_lab_text(raw)
    tsv = rows_to_tsv(rows, include_header=not args.no_header)
    write_text(args.output, tsv)


if __name__ == "__main__":
    main()
