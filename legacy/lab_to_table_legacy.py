#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
import pandas as pd
import streamlit as st
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple, TYPE_CHECKING

if TYPE_CHECKING:
    from pandas.io.formats.style import Styler

SKIP_EXACT = {"[소견]", "[의뢰의사 Comment]", "Antibiotic"}
SECTION_HEADER_RE = re.compile(r"^\s*\[진검\]\s*(.+?)\s*$")
META_LINE_RE = re.compile(r"(채혈:|접수:|보고:|IMN\()")
COLUMN_HEADER_RE = re.compile(r"^\s*검사명\s+결과값\s+단위\s+참고치\s*$")
COMMENT_LINE_RE = re.compile(r"^\s*(\[소견\]|\[의뢰의사 Comment\]|\[판독의\])")
MARKDOWN_HEADER_RE = re.compile(r"^\s*\|\s*Test\s*\|\s*Value\s*\|", re.I)
MARKDOWN_RULE_RE = re.compile(r"^\s*\|\s*-{2,}")
BULLET_PREFIX_RE = re.compile(r"^[\s\.·‥∙⋅•ㆍ]+")
EMERGENCY_PREFIX_RE = re.compile(r"^\((?:응급|응급뇨)\)\s*")
DRAW_TIME_RE = re.compile(r"채혈:\s*(\d{4}-\d{2}-\d{2})\s*(\d{2}:\d{2})")
RANGE_LIKE_NAME_RE = re.compile(r"^\s*(?:[<>≤≥]\s*\d|\d+(?:\.\d+)?\s*~\s*\d+(?:\.\d+)?)")
CONTINUATION_REF_START_RE = re.compile(r"^\s*(?:[<>≤≥]\s*\d|\d+(?:\.\d+)?\s*~\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?\s*이상|\d+(?:\.\d+)?\s*이하)")
CONTINUATION_REF_KEYWORD_RE = re.compile(r"(경계치|높음|낮음|정상|이상지질혈증|대사증후군|신부전|경도감소|고도감소)")

SECTION_TITLE_BLACKLIST = {
    "CBC with diff count & ESR", "WBC differential count", "Admission/Electro Battery(24",
    "종)", "Routine U/A with Microscope", "Gram Stain & Cul & Sensi", "Urine Microscopy",
}

QUAL_SECTION_KEYWORDS = ("미생물", "혈액은행")
REPORT_SECTION_KEYWORDS = ("영상", "소화기병센터", "내시경", "병리")
QUAL_SKIP_EXACT = {"Gram Stain & Cul & Sensi", "Blood culture", "Antibiotic"}

UNIT_ALIASES = {"㎕": "uL", "μL": "uL", "µL": "uL", "μIU": "uIU", "µIU": "uIU", "㎗": "dL"}

SPECIAL_UNIT_PATTERNS = [r"mL/min/1\.73m²", r"mOsm/kg H2O", r"mmHg", r"mm/h", r"sec", r"fL", r"pg", r"%"]
COUNT_UNIT_PATTERNS = [r"×10[⁰¹²³⁴⁵⁶⁷⁸⁹]+/(?:uL|[^\s]+)", r"×10\^?\d+/(?:uL|[^\s]+)", r"cells/HPF", r"/LPF"]
CONCENTRATION_UNIT_PATTERNS = [r"(?:mg|g|ug|ng|pg)/(?:dL|L|mL)", r"(?:IU|U|uIU)/(?:L|mL)", r"mmol/L", r"mEq/L", r"g/L", r"mg/mL"]
RATIO_UNIT_PATTERNS = [r"mmol/mol", r"mL/dL"]
UNIT_PATTERNS = SPECIAL_UNIT_PATTERNS + COUNT_UNIT_PATTERNS + CONCENTRATION_UNIT_PATTERNS + RATIO_UNIT_PATTERNS
UNIT_RE = re.compile(r"^(?:" + "|".join(f"(?:{p})" for p in UNIT_PATTERNS) + r")$", re.I)
UNIT_REF_SPLIT_RE = re.compile(r"^(?P<unit>(?:" + "|".join(f"(?:{p})" for p in UNIT_PATTERNS) + r"))\s+(?P<ref>.+)$", re.I)
FLAG_RE = re.compile(r"^[▲▼]$")

QUAL_WORDS = ["negative", "positive", "trace", "not found", "none", "some", "many", "few", "present", "absent"]
NORMAL_WORDS = ["negative", "not found", "none", "absent"]
ABNORMAL_WORDS = ["positive", "trace", "1+", "2+", "3+", "4+", "some", "many", "present"]
PLUS_PATTERN = re.compile(r"^\s*\d\+\b")
PURE_NUM_PATTERN = re.compile(r"^\s*[+-]?\d+(?:\.\d+)?\s*(?:[▲▼])?\s*$")
RANGE_NUM_PATTERN = re.compile(r"^\s*[+-]?\d+(?:\.\d+)?\s*~\s*[+-]?\d+(?:\.\d+)?")
THRESHOLD_PATTERN = re.compile(r"^\s*[+-]?\d+(?:\.\d+)?\s*(?:이상|이하)")
NUMERIC_REF_PATTERN = re.compile(r"(<|>|<=|>=|~|이상|이하|\d)")


@dataclass
class LabRow:
    name: str
    value: str
    unit: str = ""
    ref: str = ""
    raw_line: str = ""
    section: str = ""
    draw_time: str = ""
    table_title: str = ""


@dataclass
class QualRow:
    item: str
    result: str
    ref: str = ""
    status: str = "unknown"
    note: str = ""
    raw_line: str = ""
    section: str = ""
    draw_time: str = ""
    table_title: str = ""


def normalize_line(line: str) -> str:
    line = line.replace("\u3000", " ").replace("\t", " ").rstrip()
    return re.sub(r" {2,}", "  ", line)


def split_columns(line: str) -> List[str]:
    return [p.strip() for p in re.split(r" {2,}", line.strip()) if p.strip()]


def normalize_unit_token(s: str) -> str:
    s = s.strip()
    for src, dst in UNIT_ALIASES.items():
        s = s.replace(src, dst)
    superscript_map = {"0": "⁰", "1": "¹", "2": "²", "3": "³", "4": "⁴", "5": "⁵", "6": "⁶", "7": "⁷", "8": "⁸", "9": "⁹"}
    s = re.sub(r"\^(\d+)", lambda m: "".join(superscript_map.get(d, d) for d in m.group(1)), s)
    s = re.sub(r"/dl\b", "/dL", s, flags=re.I)
    s = re.sub(r"/ml\b", "/mL", s, flags=re.I)
    s = re.sub(r"/ul\b", "/uL", s, flags=re.I)
    s = re.sub(r"\bmg/dl\b", "mg/dL", s, flags=re.I)
    s = re.sub(r"\bg/dl\b", "g/dL", s, flags=re.I)
    s = re.sub(r"\bug/ml\b", "ug/mL", s, flags=re.I)
    s = re.sub(r"\bng/ml\b", "ng/mL", s, flags=re.I)
    s = re.sub(r"\bpg/ml\b", "pg/mL", s, flags=re.I)
    s = re.sub(r"\bug/dl\b", "ug/dL", s, flags=re.I)
    return s


def clean_test_name(name: str) -> str:
    """검사명에서 불필요한 기호 및 (응급) 등의 접두어 제거"""
    name = BULLET_PREFIX_RE.sub("", name)
    while True:
        new = EMERGENCY_PREFIX_RE.sub("", name)
        new = BULLET_PREFIX_RE.sub("", new)
        if new == name:
            break
        name = new
    # 연속된 공백 하나로 통합
    return re.sub(r"\s+", " ", name).strip()


def clean_section_name(section: str) -> str:
    section = re.sub(r"\[[^\]]+\]", "", section.strip())
    return re.sub(r"\s+", " ", section).strip(" -")


def extract_section_name(line: str) -> str:
    m = SECTION_HEADER_RE.match(line.strip())
    return clean_section_name(m.group(1)) if m else "Unknown"


def extract_draw_time(line: str) -> str:
    """라인에서 채혈 시각 정보를 추출 (YYYY-MM-DD HH:MM 형식)"""
    m = DRAW_TIME_RE.search(line)
    if not m:
        return ""
    return f"{m.group(1)} {m.group(2)}"


def format_title(section: str, draw_time: str) -> str:
    if draw_time:
        try:
            dt = datetime.strptime(draw_time, "%Y-%m-%d %H:%M")
            return f"{section}_{dt.year}_{dt.month:02d}_{dt.day:02d}"
        except ValueError:
            pass
    return section


def is_skip_line(line: str) -> bool:
    """파싱에서 제외해야 할 줄(스킵 대상)인지 확인"""
    s = line.strip()
    if not s:
        return True
    if s in SKIP_EXACT:
        return True
    # 헤더, 제목, 메타 정보 등 제외
    if SECTION_HEADER_RE.match(s):
        return True
    if COLUMN_HEADER_RE.match(s):
        return True
    if COMMENT_LINE_RE.match(s):
        return True
    if MARKDOWN_HEADER_RE.match(s) or MARKDOWN_RULE_RE.match(s):
        return True
    if META_LINE_RE.search(s):
        return True
    if s.startswith("- "):
        return True
    return False


def is_qualitative_section(section: str) -> bool:
    return any(k in (section or "") for k in QUAL_SECTION_KEYWORDS)


def is_report_section(section: str) -> bool:
    return any(k in (section or "") for k in REPORT_SECTION_KEYWORDS)


def is_qual_skip_line(line: str) -> bool:
    """정성 검사 파싱에서 제외할 줄인지 확인"""
    s = line.strip()
    if not s:
        return True
    if s in QUAL_SKIP_EXACT:
        return True
    if COLUMN_HEADER_RE.match(s):
        return True
    if COMMENT_LINE_RE.match(s):
        return True
    if META_LINE_RE.search(s):
        return True
    return False


def is_section_title_like(line: str) -> bool:
    return clean_test_name(line) in SECTION_TITLE_BLACKLIST


def looks_like_continuation_ref(line: str) -> bool:
    s = line.strip()
    if not s or is_skip_line(s):
        return False
    parts = split_columns(s)
    if len(parts) >= 3 and not RANGE_LIKE_NAME_RE.match(parts[0]):
        return False
    return bool(CONTINUATION_REF_START_RE.search(s) or CONTINUATION_REF_KEYWORD_RE.search(s) or s.startswith("Total cholesterol -"))


def append_ref(old_ref: str, extra: str) -> str:
    extra = re.sub(r"\s+", " ", extra).strip()
    if not old_ref:
        return extra
    return old_ref.strip() + "; " + extra


def compose_value(value: str, flag: str = "") -> str:
    """결과값과 이상 표시(▲/▼)를 합침"""
    v = value.strip()
    f = flag.strip()
    if v and f:
        return f"{v} {f}"
    return v or f


def value_flag(value: str) -> str:
    """결과값에 포함된 화살표 기호에 따라 상태 반환"""
    s_val = str(value)
    if s_val.endswith("▲"):
        return "up"
    if s_val.endswith("▼"):
        return "down"
    return ""


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\u3000", " ")).strip().lower()


def classify_row_type(result: str, ref: str = "", unit: str = "") -> str:
    r, refn, unitn = (result or "").strip().lower(), (ref or "").strip().lower(), (unit or "").strip()
    if any(w in r for w in QUAL_WORDS):
        return "qualitative"
    if PLUS_PATTERN.search(r):
        return "qualitative"
    if "cells/hpf" in r or "/lpf" in r or "cells/hpf" in refn or "/lpf" in refn:
        return "qualitative"
    if PURE_NUM_PATTERN.match(r) and unitn:
        return "quantitative"
    if PURE_NUM_PATTERN.match(r) and NUMERIC_REF_PATTERN.search(refn):
        return "quantitative"
    if RANGE_NUM_PATTERN.search(r) or THRESHOLD_PATTERN.search(r):
        return "qualitative"
    return "unknown"


def _first_num(s: str) -> Optional[float]:
    """문자열에서 첫 번째 숫자를 찾아 float로 반환 (콤마 제거 등 전처리 포함)"""
    if not s:
        return None
    # 콤마 제거 (예: 1,200 -> 1200) 및 숫자 패턴 검색
    s_clean = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s_clean)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _check_keyword_status(result: str, ref: str) -> str:
    """결과값과 참고치의 키워드 대조로 판정"""
    # 참고치에 'negative' 등이 있는데 결과에 'positive'가 있으면 abnormal
    if any(w in ref for w in NORMAL_WORDS):
        if any(w in result for w in ABNORMAL_WORDS):
            return "abnormal"
        if any(w in result for w in NORMAL_WORDS):
            return "normal"
    return "unknown"


def _check_range_status(result: str, ref: str) -> str:
    """범위 형태(10~20)의 결과/참고치 비교 판정"""
    rm = re.search(r"(-?\d+(?:\.\d+)?)\s*~\s*(-?\d+(?:\.\d+)?)", result)
    refm = re.search(r"(-?\d+(?:\.\d+)?)\s*~\s*(-?\d+(?:\.\d+)?)", ref)

    if rm and refm:
        r_low, r_high = float(rm.group(1)), float(rm.group(2))
        ref_low, ref_high = float(refm.group(1)), float(refm.group(2))
        if r_high < ref_low:
            return "low"
        if r_low > ref_high:
            return "high"
        return "normal"
    return "unknown"


def _check_threshold_status(result: str, ref: str) -> str:
    """임계치 형태(<10, >20, 이상/이하) 비교 판정"""
    # 결과가 '10 이상' 형태인 경우
    m_re = re.search(r"(-?\d+(?:\.\d+)?)\s*이상", result)
    refm = re.search(r"(-?\d+(?:\.\d+)?)\s*~\s*(-?\d+(?:\.\d+)?)", ref)
    if m_re and refm and float(m_re.group(1)) > float(refm.group(2)):
        return "high"

    m_le = re.search(r"(-?\d+(?:\.\d+)?)\s*이하", result)
    if m_le and refm and float(m_le.group(1)) < float(refm.group(1)):
        return "low"

    # 참고치가 <, >, 이상, 이하로 시작하는 경우
    rv, xv = _first_num(result), _first_num(ref)
    if rv is None or xv is None:
        return "unknown"

    if ref.startswith("<") or "이하" in ref:
        return "high" if rv > xv else "normal"
    if ref.startswith(">") or "이상" in ref:
        return "low" if rv < xv else "normal"

    return "unknown"


def classify_qual_status(result: str, ref: str = "") -> str:
    """정성/반정량 결과의 상태(normal, abnormal, high, low) 판정"""
    res_norm, ref_norm = norm_text(result), norm_text(ref)
    if not res_norm:
        return "unknown"

    # 1. 키워드 기반 판정
    status = _check_keyword_status(res_norm, ref_norm)
    if status != "unknown":
        return status

    # 2. 범위 기반 판정
    status = _check_range_status(res_norm, ref_norm)
    if status != "unknown":
        return status

    # 3. 임계치 기반 판정
    status = _check_threshold_status(res_norm, ref_norm)
    if status != "unknown":
        return status

    return "unknown"


def parse_candidate_row(line: str, section: str = "", draw_time: str = "") -> Optional[LabRow]:
    """한 줄의 텍스트가 정량 검사 결과인지 파싱"""
    line = normalize_line(line)
    if RANGE_LIKE_NAME_RE.match(line):
        return None
    
    parts = split_columns(line)
    if len(parts) < 2:
        return None

    name = value = flag = unit = ref = ""
    
    # 1. 컬럼이 4개 이상인 경우 (표준 포맷)
    if len(parts) >= 4:
        name, value = parts[0], parts[1]
        idx = 2
        
        # 이상치 화살표가 별도 컬럼인 경우
        if idx < len(parts) and FLAG_RE.match(parts[idx]):
            flag = parts[idx]
            idx += 1
        
        # 단위가 수반된 경우
        if idx < len(parts):
            pnorm = normalize_unit_token(parts[idx])
            if UNIT_RE.match(pnorm):
                unit = pnorm
                idx += 1
        
        # 나머지 부분은 참고치
        if idx < len(parts):
            ref = " ".join(parts[idx:]).strip()
        
        # 단위가 참고치 자리에 붙어 있는 경우 재추출
        if not unit and ref:
            ref_norm = normalize_unit_token(ref)
            if UNIT_RE.match(ref_norm):
                unit, ref = ref_norm, ""

    # 2. 컬럼이 3개인 경우
    elif len(parts) == 3:
        name, second, third = parts
        # 결과값에 화살표가 붙어 있는지 확인
        m = re.match(r"^(.*?)(?:\s*([▲▼]))?$", second)
        if m:
            value, flag = m.group(1).strip(), (m.group(2) or "").strip()
            
        third_norm = normalize_unit_token(third.strip())
        if UNIT_RE.match(third_norm):
            unit = third_norm
        else:
            m_ur = UNIT_REF_SPLIT_RE.match(third_norm)
            if m_ur:
                unit, ref = m_ur.group("unit").strip(), m_ur.group("ref").strip()
            else:
                ref = third.strip()

    # 3. 컬럼이 2개인 경우
    else:
        name, second = parts
        m = re.match(r"^(.*?)(?:\s*([▲▼]))?$", second)
        if m:
            value, flag = m.group(1).strip(), (m.group(2) or "").strip()

    name = clean_test_name(name)
    # 기본 검증 (검사명이나 결과값이 없으면 무시)
    if not name or name in SECTION_TITLE_BLACKLIST or name in {"검사명", "Test"} or not value:
        return None

    # 결과값 자체에 화살표가 섞여 들어간 경우 한 번 더 정리
    if not flag:
        m2 = re.match(r"^(.*?)(?:\s*([▲▼]))$", value)
        if m2:
            value, flag = m2.group(1).strip(), m2.group(2)

    return LabRow(
        name=name,
        value=compose_value(value, flag),
        unit=re.sub(r"\s+", " ", unit).strip(),
        ref=re.sub(r"\s+", " ", ref).strip(),
        raw_line=line,
        section=section,
        draw_time=draw_time,
        table_title=format_title(section or "Unknown", draw_time)
    )


def parse_qualitative_row(line: str, section: str = "", draw_time: str = "") -> Optional[QualRow]:
    """한 줄의 텍스트가 정성/반정량 검사 결과인지 파싱"""
    line = normalize_line(line)
    if not line or is_skip_line(line) or is_qual_skip_line(line) or is_section_title_like(line):
        return None

    parts = split_columns(line)
    # 항목, 결과, 참고치(있을 경우)로 분리
    if len(parts) >= 3:
        item, result, ref = clean_test_name(parts[0]), parts[1].strip(), parts[2].strip()
        note = " ".join(parts[3:]).strip() if len(parts) > 3 else ""
    elif len(parts) == 2:
        item, result, ref, note = clean_test_name(parts[0]), parts[1].strip(), "", ""
    else:
        return None

    if not item or item in {"검사명", "Test"}:
        return None

    # 해당 줄이 정성 검사형태인지 한 번 더 확인
    if classify_row_type(result, ref, "") != "qualitative":
        return None

    return QualRow(
        item=item,
        result=re.sub(r"\s+", " ", result).strip(),
        ref=re.sub(r"\s+", " ", ref).strip(),
        status=classify_qual_status(result, ref),
        note=re.sub(r"\s+", " ", note).strip(),
        raw_line=line,
        section=section,
        draw_time=draw_time,
        table_title=format_title(section or "Unknown", draw_time)
    )


def parse_lab_text(text: str) -> Tuple[List[LabRow], List[QualRow], List[str], List[str]]:
    """전체 텍스트를 순회하며 검사항목들을 추출 및 분류"""
    rows: List[LabRow] = []
    qual_rows: List[QualRow] = []
    unparsed_lines: List[str] = []
    report_lines: List[str] = []
    
    last_row: Optional[LabRow] = None
    last_qual_row: Optional[QualRow] = None
    in_comment_block = False
    current_section, current_draw_time = "Unknown", ""

    for raw in text.splitlines():
        line = normalize_line(raw)
        stripped = line.strip()
        if not stripped:
            continue

        # 1. 섹션 헤더 감지 ([진검] ...)
        if SECTION_HEADER_RE.match(stripped):
            current_section, current_draw_time = extract_section_name(stripped), ""
            in_comment_block = False
            last_row = last_qual_row = None
            continue

        # 2. 채혈/보고 시각 정보 감지
        draw_time = extract_draw_time(stripped)
        if draw_time:
            current_draw_time = draw_time
            continue

        # 3. 보고서형 섹션 (영상/병리 등)인 경우 별도 저장
        if is_report_section(current_section):
            if not is_skip_line(stripped):
                report_lines.append(stripped)
            continue

        # 4. 스킵 대상 확인 (제목 행, 코멘트 등)
        if is_skip_line(stripped):
            if stripped in {"[소견]", "[의뢰의사 Comment]"}:
                in_comment_block = True
            continue

        if in_comment_block:
            continue

        # 5. 정성 검사 섹션 처리 (미생물 등)
        if is_qualitative_section(current_section):
            qrow = parse_qualitative_row(stripped, section=current_section, draw_time=current_draw_time)
            if qrow:
                qual_rows.append(qrow)
                last_qual_row = qrow
                last_row = None
                continue
            
            # 정성 검사의 부연 설명(다음 줄) 처리
            if last_qual_row and len(split_columns(stripped)) == 1:
                extra = re.sub(r"\s+", " ", stripped).strip()
                # 참고치가 비어있으면 참고치로, 아니면 비고로 추가
                if not last_qual_row.ref:
                    last_qual_row.ref = extra
                    last_qual_row.status = classify_qual_status(last_qual_row.result, last_qual_row.ref)
                else:
                    last_qual_row.note = append_ref(last_qual_row.note, extra)
                continue
                
            if stripped not in SECTION_TITLE_BLACKLIST and not is_qual_skip_line(stripped):
                unparsed_lines.append(stripped)
            continue

        # 6. 정량 검사 처리 및 자동 타입 분류
        # 정량 검사의 참고치/단위 연장선 처리
        if last_row and looks_like_continuation_ref(stripped):
            last_row.ref = append_ref(last_row.ref, stripped)
            continue

        row = parse_candidate_row(stripped, section=current_section, draw_time=current_draw_time)
        if row:
            # 파싱된 결과가 실제로는 정성 검사일 경우 QualRow로 변환
            row_type = classify_row_type(row.value, row.ref, row.unit)
            if row_type == "qualitative":
                q_row = QualRow(
                    item=row.name,
                    result=row.value,
                    ref=row.ref,
                    status=classify_qual_status(row.value, row.ref),
                    raw_line=row.raw_line,
                    section=row.section,
                    draw_time=row.draw_time,
                    table_title=row.table_title
                )
                qual_rows.append(q_row)
                last_qual_row = q_row
                last_row = None
            else:
                rows.append(row)
                last_row = row
                last_qual_row = None
            continue

        # 7. 아무것도 해당하지 않는 줄은 미파싱 줄로 분류
        if stripped not in SECTION_TITLE_BLACKLIST:
            unparsed_lines.append(stripped)

    return rows, qual_rows, unparsed_lines, report_lines


def rows_to_tsv(rows: Iterable[LabRow], include_header: bool = True) -> str:
    lines = ["\t".join(["제목", "검사명", "결과값", "단위", "참고치"])] if include_header else []
    for r in rows:
        lines.append("\t".join([r.table_title, r.name, r.value, r.unit, r.ref]))
    return "\n".join(lines)


def rows_to_dataframe(rows: List[LabRow]) -> pd.DataFrame:
    """정량 검사 리스트를 Pandas DataFrame으로 변환"""
    import pandas as pd
    data = []
    for r in rows:
        data.append({
            "제목": r.table_title,
            "검사명": r.name,
            "결과값": r.value,
            "단위": r.unit,
            "참고치": r.ref,
            "검사종류": r.section,
            "채혈시각": r.draw_time
        })
    return pd.DataFrame(data)


def qual_rows_to_dataframe(rows: List[QualRow]) -> pd.DataFrame:
    """정성 검사 리스트를 Pandas DataFrame으로 변환"""
    import pandas as pd
    data = []
    for r in rows:
        data.append({
            "제목": r.table_title,
            "항목": r.item,
            "결과": r.result,
            "참고치": r.ref,
            "판정": r.status,
            "비고": r.note,
            "검사종류": r.section,
            "채혈시각": r.draw_time
        })
    return pd.DataFrame(data)


def report_lines_to_dataframe(lines: List[str]) -> pd.DataFrame:
    """보고서형 텍스트를 DataFrame으로 변환"""
    import pandas as pd
    return pd.DataFrame([{"보고서형 텍스트": line} for line in lines])


def rows_grouped(rows: List[LabRow]) -> Dict[str, List[LabRow]]:
    grouped: Dict[str, List[LabRow]] = {}
    for r in rows:
        grouped.setdefault(r.table_title or "Unknown", []).append(r)
    return grouped


def qual_rows_grouped(rows: List[QualRow]) -> Dict[str, List[QualRow]]:
    grouped: Dict[str, List[QualRow]] = {}
    for r in rows:
        grouped.setdefault(r.table_title or "Unknown", []).append(r)
    return grouped


def style_lab_df(df: pd.DataFrame) -> "Styler" | pd.DataFrame:
    """결과값의 ▲/▼ 방향에 따라 행 색상 지정"""
    if df.empty:
        return df

    def style_row(row):
        val = str(row.get("결과값", ""))
        flag = value_flag(val)
        if flag == "up":
            return ["color: #C62828; font-weight: bold"] * len(row)
        if flag == "down":
            return ["color: #1565C0; font-weight: bold"] * len(row)
        return [""] * len(row)

    return df.style.apply(style_row, axis=1)


def style_qual_df(df: pd.DataFrame) -> "Styler" | pd.DataFrame:
    """정성 검사의 판정 결과에 따라 행 색상 지정"""
    if df.empty:
        return df

    def style_row(row):
        status = str(row.get("판정", ""))
        if status in ("high", "abnormal"):
            return ["color: #C62828; font-weight: bold"] * len(row)
        if status == "low":
            return ["color: #1565C0; font-weight: bold"] * len(row)
        return [""] * len(row)

    return df.style.apply(style_row, axis=1)


def build_excel_bytes(rows: List[LabRow], qual_rows: List[QualRow], raw_text: str, unparsed_lines: List[str], report_lines: List[str]) -> BytesIO:
    """추출된 데이터를 엑셀 파일(다중 시트)로 변환"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 1. 시트별 데이터프레임 구성
    parsed_df = rows_to_dataframe(rows)
    qual_df = qual_rows_to_dataframe(qual_rows)
    raw_df = pd.DataFrame({"원문": raw_text.splitlines()})
    unparsed_df = pd.DataFrame({"파싱되지 않은 줄": unparsed_lines})
    report_df = report_lines_to_dataframe(report_lines)
    
    grouped = rows_grouped(rows)
    qual_grouped = qual_rows_grouped(qual_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 데이터 시트들 저장
        parsed_df.to_excel(writer, index=False, sheet_name="Parsed_Labs")
        qual_df.to_excel(writer, index=False, sheet_name="Qualitative_Results")
        raw_df.to_excel(writer, index=False, sheet_name="Original_Text")
        unparsed_df.to_excel(writer, index=False, sheet_name="Unparsed")
        report_df.to_excel(writer, index=False, sheet_name="Ignored_Reports")
        
        wb = writer.book
        # 그룹화 시트를 맨 앞에 생성
        ws_group = wb.create_sheet("Grouped_Tables", 0)
        ws_qgroup = wb.create_sheet("Grouped_Qualitative", 1)

        # 스타일 정의
        title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        red_font = Font(color="C62828", bold=True)
        blue_font = Font(color="1565C0", bold=True)

        # --- Grouped_Tables 시트 작성 (정량) ---
        cur = 1
        for title, group_rows in grouped.items():
            # 제목행 (제목 합치기 및 스타일)
            ws_group.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
            cell = ws_group.cell(row=cur, column=1, value=title)
            cell.font, cell.fill, cell.alignment = title_font, title_fill, Alignment(horizontal="left")
            cur += 1

            # 헤더행
            for i, h in enumerate(["검사명", "결과값", "단위", "참고치"], start=1):
                cc = ws_group.cell(row=cur, column=i, value=h)
                cc.font, cc.fill = header_font, header_fill
            cur += 1

            # 데이터행
            for r in group_rows:
                cells = [
                    ws_group.cell(row=cur, column=1, value=r.name),
                    ws_group.cell(row=cur, column=2, value=r.value),
                    ws_group.cell(row=cur, column=3, value=r.unit),
                    ws_group.cell(row=cur, column=4, value=r.ref)
                ]
                vf = value_flag(r.value)
                if vf == "up":
                    for cell in cells: cell.font = red_font
                elif vf == "down":
                    for cell in cells: cell.font = blue_font
                cur += 1
            cur += 2

        # --- Grouped_Qualitative 시트 작성 (정성) ---
        cur = 1
        for title, group_rows in qual_grouped.items():
            ws_qgroup.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=5)
            cell = ws_qgroup.cell(row=cur, column=1, value=title)
            cell.font, cell.fill, cell.alignment = title_font, title_fill, Alignment(horizontal="left")
            cur += 1

            for i, h in enumerate(["항목", "결과", "참고치", "판정", "비고"], start=1):
                cc = ws_qgroup.cell(row=cur, column=i, value=h)
                cc.font, cc.fill = header_font, header_fill
            cur += 1

            for r in group_rows:
                cells = [
                    ws_qgroup.cell(row=cur, column=1, value=r.item),
                    ws_qgroup.cell(row=cur, column=2, value=r.result),
                    ws_qgroup.cell(row=cur, column=3, value=r.ref),
                    ws_qgroup.cell(row=cur, column=4, value=r.status),
                    ws_qgroup.cell(row=cur, column=5, value=r.note)
                ]
                if r.status in ("high", "abnormal"):
                    for cell in cells: cell.font = red_font
                elif r.status == "low":
                    for cell in cells: cell.font = blue_font
                cur += 1
            cur += 2

        # 컬럼 너비 및 전체 스타일 정리
        sheets_to_auto = {
            "Parsed_Labs": (parsed_df, 3), # 결과값 컬럼 인덱스
            "Qualitative_Results": (qual_df, None), 
            "Original_Text": (raw_df, None), 
            "Unparsed": (unparsed_df, None), 
            "Ignored_Reports": (report_df, None)
        }

        for sheet_name, (df, val_col_idx) in sheets_to_auto.items():
            ws = writer.sheets[sheet_name]
            # 너비 자동 조정
            for idx, col in enumerate(df.columns, start=1):
                max_len = max([len(str(col))] + [len(str(v)) for v in df[col].fillna("")]) if not df.empty else len(str(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)

            # 데이터 시트에서의 색상 적용 (Parsed_Labs)
            if sheet_name == "Parsed_Labs" and not df.empty:
                for row_idx in range(2, len(df) + 2):
                    val = str(ws.cell(row=row_idx, column=3).value or "") # 3번 컬럼: 결과값
                    vf = value_flag(val)
                    if vf:
                        font = red_font if vf == "up" else blue_font
                        for col_idx in range(1, 8):
                            ws.cell(row=row_idx, column=col_idx).font = font
            
            # 데이터 시트에서의 색상 적용 (Qualitative_Results)
            if sheet_name == "Qualitative_Results" and not df.empty:
                # '판정' 컬럼 찾기
                status_col = next((i for i in range(1, ws.max_column + 1) if ws.cell(row=1, column=i).value == "판정"), None)
                if status_col:
                    for row_idx in range(2, len(df) + 2):
                        status = str(ws.cell(row=row_idx, column=status_col).value or "")
                        font = red_font if status in ("high", "abnormal") else blue_font if status == "low" else None
                        if font:
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = font

        output.seek(0)
    return output


def render_lab_to_excel_tool() -> None:
    """Streamlit 도구 화면 렌더링"""
    st.title("🧪 Lab → Excel 변환기")
    st.caption("정량/정성 검사를 표로 정리한 뒤, 엑셀 파일로 다운로드합니다.")
    st.info(
        "보고서형 검사(영상/내시경/병리)는 표 변환 대신 Text Cleaner 사용을 권장합니다. "
        "정량은 기존 방식대로, 정성은 결과/참고치 비교로 판정합니다."
    )
    
    default_sample = (
        "[진검]  응급혈액[WB, EDTA]\n"
        "　채혈: 2026-04-05 17:48  접수: 2026-04-06 13:45  IMN(조영일)  보고: 2026-04-06 14:04  -\n"
        "　검사명                               결과값       단위         참고치\n"
        "　　(응급)WBC                          9.13         ×10³/㎕    4~10\n"
        "　　(응급)RBC                          3.27  ▼     ×10^6/㎕    4.2~6.3\n"
        "　　(응급)Protein                      2+ (65~200mg/dl)          Negative\n"
    )

    with st.expander("예시 데이터 넣기", expanded=False):
        if st.button("샘플 입력", use_container_width=True):
            st.session_state["lab_excel_input"] = default_sample
            st.rerun()

    raw_text = st.text_area(
        "Lab / EMR 원문 붙여넣기", 
        key="lab_excel_input", 
        height=360, 
        placeholder="여기에 검사 결과 원문을 붙여넣으세요..."
    )
    
    c1, c2 = st.columns(2)
    run = c1.button("표 만들기", type="primary", use_container_width=True)
    clear = c2.button("입력 지우기", use_container_width=True)
    
    if clear:
        st.session_state["lab_excel_input"] = ""
        st.session_state["lab_results"] = None
        st.rerun()

    # --- 데이터 처리 로직 ---
    if run:
        if not raw_text.strip():
            st.warning("원문을 먼저 붙여넣어 주세요.")
            st.session_state["lab_results"] = None
        else:
            rows, qual_rows, unparsed_lines, report_lines = parse_lab_text(raw_text)
            
            # 파싱 결과를 세션 상태에 저장
            st.session_state["lab_results"] = {
                "rows": rows,
                "qual_rows": qual_rows,
                "unparsed_lines": unparsed_lines,
                "report_lines": report_lines,
                "df": rows_to_dataframe(rows),
                "qual_df": qual_rows_to_dataframe(qual_rows),
                "excel_bytes": build_excel_bytes(rows, qual_rows, raw_text, unparsed_lines, report_lines),
                "tsv_text": rows_to_tsv(rows),
                "grouped": rows_grouped(rows),
                "qual_grouped": qual_rows_grouped(qual_rows),
            }

    # --- 결과 렌더링 로직 ---
    if st.session_state.get("lab_results") is not None:
        res = st.session_state["lab_results"]
        rows, qual_rows = res["rows"], res["qual_rows"]
        unparsed_lines, report_lines = res["unparsed_lines"], res["report_lines"]
        df, qual_df = res["df"], res["qual_df"]
        grouped, qual_grouped = res["grouped"], res["qual_grouped"]

        st.success(f"정량 {len(rows)}개 row, 정성 {len(qual_rows)}개 row를 추출했습니다.")
        
        # 메트릭 대시보드
        cols = st.columns(6)
        cols[0].metric("정량 row", len(rows))
        cols[1].metric("정성 row", len(qual_rows))
        cols[2].metric("정량 표 수", len(grouped))
        cols[3].metric("미파싱 줄", len(unparsed_lines))
        cols[4].metric("보고서 줄", len(report_lines))
        cols[5].metric("원문 줄 수", len(raw_text.splitlines()))

        # 다운로드 버튼
        now_str = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d_%H%M")
        dl_c1, dl_c2 = st.columns(2)
        dl_c1.download_button(
            "📥 엑셀 다운로드", 
            data=res["excel_bytes"], 
            file_name=f"table_{now_str}.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )
        dl_c2.download_button(
            "📥 TSV 다운로드", 
            data=res["tsv_text"], 
            file_name=f"table_{now_str}.tsv", 
            mime="text/tab-separated-values", 
            use_container_width=True
        )

        st.markdown("---")
        
        # 정량 결과 미리보기
        st.markdown("### 📊 정량 결과 미리보기")
        for title, group_rows in grouped.items():
            st.markdown(f"#### {title}")
            display_df = rows_to_dataframe(group_rows)[["검사명", "결과값", "단위", "참고치"]]
            st.dataframe(style_lab_df(display_df), use_container_width=True, hide_index=True)
            st.write("")

        # 정성 결과 미리보기
        if qual_grouped:
            st.markdown("### 📝 정성 결과 미리보기")
            for title, group_rows in qual_grouped.items():
                st.markdown(f"#### {title}")
                display_qdf = qual_rows_to_dataframe(group_rows)[["항목", "결과", "참고치", "판정", "비고"]]
                st.dataframe(style_qual_df(display_qdf), use_container_width=True, hide_index=True)
                st.write("")

        # 전체 데이터 익스팬더
        with st.expander("🔍 전체 정량 row 테이블 보기", expanded=False):
            st.dataframe(style_lab_df(df), use_container_width=True, hide_index=True)

        if not qual_df.empty:
            with st.expander("🔍 전체 정성 row 테이블 보기", expanded=False):
                st.dataframe(style_qual_df(qual_df), use_container_width=True, hide_index=True)

        if report_lines:
            st.info("💡 보고서형 검사(영상/내시경/병리 등)는 텍스트 클리너 사용을 권장합니다.")
            with st.expander("📄 텍스트 클리너로 넘길 보고서형 줄 보기", expanded=False):
                st.code("\n".join(report_lines), language="text")

        if unparsed_lines:
            with st.expander("⚠️ 파싱되지 않은 줄 보기", expanded=False):
                st.code("\n".join(unparsed_lines), language="text")


def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Raw EMR/lab text -> Excel-friendly TSV")
    p.add_argument("input", nargs="?", help="input text file path. If omitted, read from stdin.")
    p.add_argument("-o", "--output", help="output TSV file path. If omitted, print to stdout.")
    p.add_argument("--no-header", action="store_true", help="omit TSV header row")
    return p


def read_text(path: Optional[str]) -> str:
    if path:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    return sys.stdin.read()


def write_text(path: Optional[str], text: str) -> None:
    if path:
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write(text)
    else:
        sys.stdout.write(text)


def main() -> None:
    args = build_argparser().parse_args()
    raw = read_text(args.input)
    rows, _, _, _ = parse_lab_text(raw)
    tsv = rows_to_tsv(rows, include_header=not args.no_header)
    write_text(args.output, tsv)


if __name__ == "__main__":
    main()
