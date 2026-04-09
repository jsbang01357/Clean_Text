import pandas as pd
from io import BytesIO
from typing import Dict, Iterable, List, Optional, TYPE_CHECKING, Union
from core.lab_parser import LabRow, QualRow, value_flag

if TYPE_CHECKING:
    from pandas.io.formats.style import Styler

def rows_to_tsv(rows: Iterable[LabRow], include_header: bool = True) -> str:
    lines = ["\t".join(["제목", "검사명", "결과값", "단위", "참고치"])] if include_header else []
    for r in rows:
        lines.append("\t".join([r.table_title, r.name, r.value, r.unit, r.ref]))
    return "\n".join(lines)

def rows_to_dataframe(rows: List[LabRow]) -> pd.DataFrame:
    data = []
    for r in rows:
        data.append({
            "제목": r.table_title,
            "검사명": r.name,
            "결과값": r.value,
            "단위": r.unit,
            "참고치": r.ref,
            "검사종류": r.section,
            "채혈시각": r.draw_time
        })
    return pd.DataFrame(data)

def qual_rows_to_dataframe(rows: List[QualRow]) -> pd.DataFrame:
    data = []
    for r in rows:
        data.append({
            "제목": r.table_title,
            "항목": r.item,
            "결과": r.result,
            "참고치": r.ref,
            "판정": r.status,
            "비고": r.note,
            "검사종류": r.section,
            "채혈시각": r.draw_time
        })
    return pd.DataFrame(data)

def report_lines_to_dataframe(lines: List[str]) -> pd.DataFrame:
    return pd.DataFrame([{"보고서형 텍스트": line} for line in lines])

def rows_grouped(rows: List[LabRow]) -> Dict[str, List[LabRow]]:
    grouped: Dict[str, List[LabRow]] = {}
    for r in rows:
        grouped.setdefault(r.table_title or "Unknown", []).append(r)
    return grouped

def qual_rows_grouped(rows: List[QualRow]) -> Dict[str, List[QualRow]]:
    grouped: Dict[str, List[QualRow]] = {}
    for r in rows:
        grouped.setdefault(r.table_title or "Unknown", []).append(r)
    return grouped

def style_lab_df(df: pd.DataFrame) -> Union["Styler", pd.DataFrame]:
    """결과값의 ▲/▼ 방향에 따라 행 색상 지정"""
    if df.empty:
        return df

    def style_row(row):
        val = str(row.get("결과값", ""))
        flag = value_flag(val)
        if flag == "up":
            return ["color: #C62828; font-weight: bold"] * len(row)
        if flag == "down":
            return ["color: #1565C0; font-weight: bold"] * len(row)
        return [""] * len(row)

    return df.style.apply(style_row, axis=1)

def style_qual_df(df: pd.DataFrame) -> Union["Styler", pd.DataFrame]:
    """정성 검사의 판정 결과에 따라 행 색상 지정"""
    if df.empty:
        return df

    def style_row(row):
        status = str(row.get("판정", ""))
        if status in ("high", "abnormal"):
            return ["color: #C62828; font-weight: bold"] * len(row)
        if status == "low":
            return ["color: #1565C0; font-weight: bold"] * len(row)
        return [""] * len(row)

    return df.style.apply(style_row, axis=1)

def build_excel_bytes(rows: List[LabRow], qual_rows: List[QualRow], raw_text: str, unparsed_lines: List[str], report_lines: List[str]) -> BytesIO:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    parsed_df = rows_to_dataframe(rows)
    qual_df = qual_rows_to_dataframe(qual_rows)
    raw_df = pd.DataFrame({"원문": raw_text.splitlines()})
    unparsed_df = pd.DataFrame({"파싱되지 않은 줄": unparsed_lines})
    report_df = report_lines_to_dataframe(report_lines)
    
    grouped = rows_grouped(rows)
    qual_grouped = qual_rows_grouped(qual_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        parsed_df.to_excel(writer, index=False, sheet_name="Parsed_Labs")
        qual_df.to_excel(writer, index=False, sheet_name="Qualitative_Results")
        raw_df.to_excel(writer, index=False, sheet_name="Original_Text")
        unparsed_df.to_excel(writer, index=False, sheet_name="Unparsed")
        report_df.to_excel(writer, index=False, sheet_name="Ignored_Reports")
        
        wb = writer.book
        ws_group = wb.create_sheet("Grouped_Tables", 0)
        ws_qgroup = wb.create_sheet("Grouped_Qualitative", 1)

        title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        red_font = Font(color="C62828", bold=True)
        blue_font = Font(color="1565C0", bold=True)

        cur = 1
        for title, group_rows in grouped.items():
            ws_group.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
            cell = ws_group.cell(row=cur, column=1, value=title)
            cell.font, cell.fill, cell.alignment = title_font, title_fill, Alignment(horizontal="left")
            cur += 1

            for i, h in enumerate(["검사명", "결과값", "단위", "참고치"], start=1):
                cc = ws_group.cell(row=cur, column=i, value=h)
                cc.font, cc.fill = header_font, header_fill
            cur += 1

            for r in group_rows:
                cells = [
                    ws_group.cell(row=cur, column=1, value=r.name),
                    ws_group.cell(row=cur, column=2, value=r.value),
                    ws_group.cell(row=cur, column=3, value=r.unit),
                    ws_group.cell(row=cur, column=4, value=r.ref)
                ]
                vf = value_flag(r.value)
                if vf == "up":
                    for cell in cells: cell.font = red_font
                elif vf == "down":
                    for cell in cells: cell.font = blue_font
                cur += 1
            cur += 2

        cur = 1
        for title, group_rows in qual_grouped.items():
            ws_qgroup.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=5)
            cell = ws_qgroup.cell(row=cur, column=1, value=title)
            cell.font, cell.fill, cell.alignment = title_font, title_fill, Alignment(horizontal="left")
            cur += 1

            for i, h in enumerate(["항목", "결과", "참고치", "판정", "비고"], start=1):
                cc = ws_qgroup.cell(row=cur, column=i, value=h)
                cc.font, cc.fill = header_font, header_fill
            cur += 1

            for r in group_rows:
                cells = [
                    ws_qgroup.cell(row=cur, column=1, value=r.item),
                    ws_qgroup.cell(row=cur, column=2, value=r.result),
                    ws_qgroup.cell(row=cur, column=3, value=r.ref),
                    ws_qgroup.cell(row=cur, column=4, value=r.status),
                    ws_qgroup.cell(row=cur, column=5, value=r.note)
                ]
                if r.status in ("high", "abnormal"):
                    for cell in cells: cell.font = red_font
                elif r.status == "low":
                    for cell in cells: cell.font = blue_font
                cur += 1
            cur += 2

        sheets_to_auto = {
            "Parsed_Labs": (parsed_df, 3), 
            "Qualitative_Results": (qual_df, None), 
            "Original_Text": (raw_df, None), 
            "Unparsed": (unparsed_df, None), 
            "Ignored_Reports": (report_df, None)
        }

        for sheet_name, (df, val_col_idx) in sheets_to_auto.items():
            ws = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, start=1):
                max_len = max([len(str(col))] + [len(str(v)) for v in df[col].fillna("")]) if not df.empty else len(str(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)

            if sheet_name == "Parsed_Labs" and not df.empty:
                for row_idx in range(2, len(df) + 2):
                    val = str(ws.cell(row=row_idx, column=3).value or "") 
                    vf = value_flag(val)
                    if vf:
                        font = red_font if vf == "up" else blue_font
                        for col_idx in range(1, 8):
                            ws.cell(row=row_idx, column=col_idx).font = font
            
            if sheet_name == "Qualitative_Results" and not df.empty:
                status_col = next((i for i in range(1, ws.max_column + 1) if ws.cell(row=1, column=i).value == "판정"), None)
                if status_col:
                    for row_idx in range(2, len(df) + 2):
                        status = str(ws.cell(row=row_idx, column=status_col).value or "")
                        font = red_font if status in ("high", "abnormal") else blue_font if status == "low" else None
                        if font:
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).font = font

    output.seek(0)
    return output
